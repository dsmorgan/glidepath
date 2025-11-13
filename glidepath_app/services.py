import csv
import io
import os
import re
from decimal import Decimal, ROUND_HALF_UP
from typing import Dict, Optional

import openpyxl
from django.db import transaction

from .models import (
    AssetCategory,
    AssetClass,
    CategoryAllocation,
    ClassAllocation,
    GlidepathRule,
    RuleSet,
    AssumptionUpload,
    AssumptionData,
    User,
)

ASSET_CLASSES = ["Stocks", "Bonds", "Crypto", "Other"]


def _parse_percent(value: str) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    value = value.strip()
    if value.endswith("%"):
        value = value[:-1]
    return Decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _normalize_header(header: str) -> str:
    return re.sub(r"\s*:\s*", ":", header).strip()


def _unique_ruleset_name(base: str) -> str:
    existing = set(RuleSet.objects.values_list("name", flat=True))
    name = base
    idx = 1
    while name in existing:
        name = f"{base} ({idx})"
        idx += 1
    return name


def import_glidepath_rules(file_obj) -> RuleSet:
    name = os.path.splitext(os.path.basename(getattr(file_obj, "name", "") or "rules"))[0]
    name = name or "rules"
    ruleset = RuleSet.objects.create(name=_unique_ruleset_name(name))

    text = io.TextIOWrapper(file_obj, encoding="utf-8")
    reader = csv.DictReader(text)
    if reader.fieldnames:
        reader.fieldnames = [_normalize_header(c) for c in reader.fieldnames]
    required = {"gt-retire-age", "lt-retire-age"}
    if not required.issubset(reader.fieldnames or []):
        raise ValueError("Missing required columns: gt-retire-age and lt-retire-age")

    class_cols = [c for c in reader.fieldnames if c in ASSET_CLASSES]
    category_cols = [c for c in reader.fieldnames if ":" in c]

    rows = []
    for row in reader:
        gt = int(row["gt-retire-age"])
        lt = int(row["lt-retire-age"])
        gt = max(-100, gt)
        lt = min(100, lt)
        if gt >= lt:
            raise ValueError("gt-retire-age must be less than lt-retire-age")
        rows.append((gt, lt, row))

    rows.sort(key=lambda x: x[0])
    current = -100
    for gt, lt, _ in rows:
        if gt > current:
            raise ValueError(f"Missing rules for ages {current} to {gt}")
        if gt < current:
            raise ValueError(f"Overlapping rules for ages {gt} to {current}")
        current = lt
    if current != 100:
        raise ValueError(f"Missing rules for ages {current} to 100")

    with transaction.atomic():
        for gt, lt, row in rows:
            rule = GlidepathRule.objects.create(
                ruleset=ruleset, gt_retire_age=gt, lt_retire_age=lt
            )

            # class allocations
            class_pcts: Dict[str, Decimal] = {}
            total_class = Decimal("0")
            for col in class_cols:
                pct = _parse_percent(row.get(col, "0"))
                if pct:
                    class_pcts[col] = pct
                    total_class += pct
            if total_class > Decimal("100"):
                raise ValueError("Class allocations exceed 100%")
            if "Other" not in class_pcts:
                class_pcts["Other"] = (Decimal("100") - total_class).quantize(
                    Decimal("0.01")
                )
            total_class = sum(class_pcts.values())
            if total_class != Decimal("100"):
                raise ValueError("Class allocations must total 100%")

            # categories
            category_totals: Dict[str, Dict[str, Decimal]] = {}
            total_category = Decimal("0")
            for col in category_cols:
                cls, cat = col.split(":", 1)
                if cls not in ASSET_CLASSES:
                    raise ValueError(f"Invalid asset class '{cls}' in column '{col}'")
                pct = _parse_percent(row.get(col, "0"))
                if pct:
                    category_totals.setdefault(cls, {})[cat] = pct
                    total_category += pct
            if total_category != Decimal("100"):
                raise ValueError("Category allocations must total 100%")

            for cls, cats in category_totals.items():
                cat_sum = sum(cats.values())
                cls_pct = class_pcts.get(cls, Decimal("0"))
                if cls_pct != cat_sum:
                    raise ValueError(
                        f"Category allocations for {cls} ({cat_sum}%) do not match class allocation ({cls_pct}%)"
                    )

            # ensure asset classes exist
            existing_classes = {
                ac.name: ac for ac in AssetClass.objects.filter(name__in=ASSET_CLASSES)
            }
            for name in ASSET_CLASSES:
                if name not in existing_classes:
                    existing_classes[name] = AssetClass.objects.create(name=name)

            for cls_name, pct in class_pcts.items():
                ClassAllocation.objects.create(
                    rule=rule, asset_class=existing_classes[cls_name], percentage=pct
                )

            for cls_name, cats in category_totals.items():
                for cat_name, pct in cats.items():
                    category, _ = AssetCategory.objects.get_or_create(
                        name=cat_name, asset_class=existing_classes[cls_name]
                    )
                    CategoryAllocation.objects.create(
                        rule=rule, asset_category=category, percentage=pct
                    )

    return ruleset


def export_glidepath_rules(ruleset: RuleSet) -> str:
    asset_classes = list(
        AssetClass.objects.filter(
            classallocation__rule__ruleset=ruleset,
            classallocation__percentage__gt=0,
        )
        .distinct()
        .order_by("name")
    )
    categories = list(
        AssetCategory.objects.filter(
            categoryallocation__rule__ruleset=ruleset,
            categoryallocation__percentage__gt=0,
        )
        .distinct()
        .order_by("asset_class__name", "name")
    )

    headers = ["gt-retire-age", "lt-retire-age"]
    headers += [ac.name for ac in asset_classes]
    headers += [f"{c.asset_class.name}:{c.name}" for c in categories]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    for rule in GlidepathRule.objects.filter(ruleset=ruleset).order_by(
        "gt_retire_age"
    ):
        row = {
            "gt-retire-age": rule.gt_retire_age,
            "lt-retire-age": rule.lt_retire_age,
        }
        class_map = {
            ca.asset_class.name: ca.percentage for ca in rule.class_allocations.all()
        }
        category_map = {
            f"{ca.asset_category.asset_class.name}:{ca.asset_category.name}": ca.percentage
            for ca in rule.category_allocations.all()
        }
        for ac in asset_classes:
            row[ac.name] = f"{class_map.get(ac.name, Decimal('0'))}%"
        for cat in categories:
            key = f"{cat.asset_class.name}:{cat.name}"
            row[key] = f"{category_map.get(key, Decimal('0'))}%"
        writer.writerow(row)
    return output.getvalue()


def _parse_decimal(value) -> Optional[Decimal]:
    """Parse a cell value to Decimal, handling None and various formats."""
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except:
        return None


def import_blackrock_assumptions(file_obj, user: User) -> AssumptionUpload:
    """
    Parse a BlackRock Capital Market Assumptions XLSX file.

    Expected structure:
    - Row 1: Title
    - Row 2: File date/time in column A, grouped column headers (e.g., "Expected returns")
    - Row 3: Specific column names (e.g., "5 year", "7 year", etc.)
    - Row 4+: Data rows (filter for Currency == "USD")

    If a file with the same file_datetime already exists, it will be replaced.
    """
    filename = os.path.basename(getattr(file_obj, "name", "") or "assumptions.xlsx")

    # Load the workbook
    wb = openpyxl.load_workbook(file_obj)
    sheet = wb[wb.sheetnames[0]]  # First sheet only

    # Extract file date/time from A2
    file_datetime = str(sheet.cell(row=2, column=1).value or "").strip()
    if not file_datetime:
        raise ValueError("Could not extract date/time from file (Cell A2)")

    # Parse headers from rows 2 and 3
    # Build a mapping of column index to field name
    col_mapping = {}

    # First, get Row 2 grouped headers
    row2_groups = {}
    current_group = ""
    for j in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=2, column=j)
        if cell.value:
            current_group = str(cell.value).strip()
        row2_groups[j] = current_group

    # Then combine with Row 3 specific columns
    for j in range(1, sheet.max_column + 1):
        cell3 = sheet.cell(row=3, column=j)
        if cell3.value:
            col_name = str(cell3.value).strip()
            group = row2_groups.get(j, "")

            # Special handling for basic columns
            if col_name == "Currency":
                col_mapping[j] = ("currency", None)
            elif col_name == "Asset class":
                col_mapping[j] = ("asset_class", None)
            elif col_name == "Asset":
                col_mapping[j] = ("asset", None)
            elif col_name == "Index":
                col_mapping[j] = ("index", None)
            elif group == "Expected returns":
                col_mapping[j] = ("expected_return", col_name)
            elif group == "Lower interquartile range (25th percentile)":
                col_mapping[j] = ("lower_iqr", col_name)
            elif group == "Upper interquartile range (25th percentile)":
                col_mapping[j] = ("upper_iqr", col_name)
            elif group == "Lower mean uncertainty":
                col_mapping[j] = ("lower_uncertainty", col_name)
            elif group == "Upper mean uncertainty":
                col_mapping[j] = ("upper_uncertainty", col_name)
            elif group == "Volatility" and not col_name:
                col_mapping[j] = ("volatility", None)
            elif group == "Correlation":
                if col_name == "Government bonds":
                    col_mapping[j] = ("correlation_govt_bonds", None)
                elif col_name == "Equities":
                    col_mapping[j] = ("correlation_equities", None)

    # Map year names to field suffixes
    year_mapping = {
        "5 year": "5yr",
        "7 year": "7yr",
        "10 year": "10yr",
        "15 year": "15yr",
        "20 year": "20yr",
        "25 year": "25yr",
        "30 year": "30yr",
    }

    # Parse data rows (starting from row 4)
    data_rows = []
    for i in range(4, sheet.max_row + 1):
        # Check if this is a USD row
        currency_col = None
        for col_idx, (field_type, _) in col_mapping.items():
            if field_type == "currency":
                currency_col = col_idx
                break

        if currency_col is None:
            continue

        currency = sheet.cell(row=i, column=currency_col).value
        if str(currency).strip().upper() != "USD":
            continue

        # Extract all values for this row
        row_data = {
            "currency": None,
            "asset_class": None,
            "asset": None,
            "index": None,
        }

        for col_idx, (field_type, year_name) in col_mapping.items():
            cell_value = sheet.cell(row=i, column=col_idx).value

            if field_type in ["currency", "asset_class", "asset", "index"]:
                row_data[field_type] = str(cell_value or "").strip()
            elif field_type == "volatility":
                row_data["volatility"] = _parse_decimal(cell_value)
            elif field_type == "correlation_govt_bonds":
                row_data["correlation_govt_bonds"] = _parse_decimal(cell_value)
            elif field_type == "correlation_equities":
                row_data["correlation_equities"] = _parse_decimal(cell_value)
            elif year_name and year_name in year_mapping:
                suffix = year_mapping[year_name]
                field_name = f"{field_type}_{suffix}"
                row_data[field_name] = _parse_decimal(cell_value)

        # Only add row if it has at least an asset
        if row_data.get("asset"):
            data_rows.append(row_data)

    if not data_rows:
        raise ValueError("No USD data rows found in file")

    # Check if upload with this file_datetime already exists
    with transaction.atomic():
        existing_upload = AssumptionUpload.objects.filter(file_datetime=file_datetime).first()
        if existing_upload:
            # Delete existing upload (cascade will delete data rows)
            existing_upload.delete()

        # Create new upload
        upload = AssumptionUpload.objects.create(
            user=user,
            file_datetime=file_datetime,
            upload_type="blackrock",
            filename=filename,
            entry_count=len(data_rows),
        )

        # Create data rows
        for row_data in data_rows:
            AssumptionData.objects.create(
                upload=upload,
                currency=row_data.get("currency", ""),
                asset_class=row_data.get("asset_class", ""),
                asset=row_data.get("asset", ""),
                index=row_data.get("index", ""),
                expected_return_5yr=row_data.get("expected_return_5yr"),
                expected_return_7yr=row_data.get("expected_return_7yr"),
                expected_return_10yr=row_data.get("expected_return_10yr"),
                expected_return_15yr=row_data.get("expected_return_15yr"),
                expected_return_20yr=row_data.get("expected_return_20yr"),
                expected_return_25yr=row_data.get("expected_return_25yr"),
                expected_return_30yr=row_data.get("expected_return_30yr"),
                lower_iqr_5yr=row_data.get("lower_iqr_5yr"),
                lower_iqr_7yr=row_data.get("lower_iqr_7yr"),
                lower_iqr_10yr=row_data.get("lower_iqr_10yr"),
                lower_iqr_15yr=row_data.get("lower_iqr_15yr"),
                lower_iqr_20yr=row_data.get("lower_iqr_20yr"),
                lower_iqr_25yr=row_data.get("lower_iqr_25yr"),
                lower_iqr_30yr=row_data.get("lower_iqr_30yr"),
                upper_iqr_5yr=row_data.get("upper_iqr_5yr"),
                upper_iqr_7yr=row_data.get("upper_iqr_7yr"),
                upper_iqr_10yr=row_data.get("upper_iqr_10yr"),
                upper_iqr_15yr=row_data.get("upper_iqr_15yr"),
                upper_iqr_20yr=row_data.get("upper_iqr_20yr"),
                upper_iqr_25yr=row_data.get("upper_iqr_25yr"),
                upper_iqr_30yr=row_data.get("upper_iqr_30yr"),
                lower_uncertainty_5yr=row_data.get("lower_uncertainty_5yr"),
                lower_uncertainty_7yr=row_data.get("lower_uncertainty_7yr"),
                lower_uncertainty_10yr=row_data.get("lower_uncertainty_10yr"),
                lower_uncertainty_15yr=row_data.get("lower_uncertainty_15yr"),
                lower_uncertainty_20yr=row_data.get("lower_uncertainty_20yr"),
                lower_uncertainty_25yr=row_data.get("lower_uncertainty_25yr"),
                lower_uncertainty_30yr=row_data.get("lower_uncertainty_30yr"),
                upper_uncertainty_5yr=row_data.get("upper_uncertainty_5yr"),
                upper_uncertainty_7yr=row_data.get("upper_uncertainty_7yr"),
                upper_uncertainty_10yr=row_data.get("upper_uncertainty_10yr"),
                upper_uncertainty_15yr=row_data.get("upper_uncertainty_15yr"),
                upper_uncertainty_20yr=row_data.get("upper_uncertainty_20yr"),
                upper_uncertainty_25yr=row_data.get("upper_uncertainty_25yr"),
                upper_uncertainty_30yr=row_data.get("upper_uncertainty_30yr"),
                volatility=row_data.get("volatility"),
                correlation_govt_bonds=row_data.get("correlation_govt_bonds"),
                correlation_equities=row_data.get("correlation_equities"),
            )

    return upload
